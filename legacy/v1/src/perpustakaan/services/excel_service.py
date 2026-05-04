"""Import/export data dari/ke file Excel (.xlsx).

Mendukung template impor anggota & buku sesuai SIM-Perpus asli.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from perpustakaan.config import EXPORTS_DIR
from perpustakaan.db.connection import Database, get_db
from perpustakaan.models import anggota as anggota_repo
from perpustakaan.models import buku as buku_repo

# ---------------------------------------------------------------------------
# Template generators
# ---------------------------------------------------------------------------
ANGGOTA_HEADERS: list[str] = [
    "kode_anggota",
    "nama",
    "jenis_kelamin",
    "kelas",
    "jurusan",
    "tempat_lahir",
    "tanggal_lahir",
    "no_telp",
    "email",
    "alamat",
    "catatan",
]

BUKU_HEADERS: list[str] = [
    "kode_buku",
    "judul",
    "pengarang",
    "penerbit",
    "tahun_terbit",
    "kode_ddc",
    "kategori",
    "isbn",
    "jumlah_eksemplar",
    "sumber",
    "harga",
    "bahasa",
    "rak",
    "deskripsi",
]


def template_anggota(path: Path | None = None) -> Path:
    out = path or EXPORTS_DIR / "Template_Impor_Anggota.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Anggota"
    ws.append(ANGGOTA_HEADERS)
    ws.append(
        [
            "A0001",
            "Contoh Nama",
            "L",
            "VII A",
            "-",
            "Jakarta",
            "2010-01-01",
            "0812xxxx",
            "user@example.com",
            "Jl. Contoh",
            "",
        ]
    )
    wb.save(out)
    return out


def template_buku(path: Path | None = None) -> Path:
    out = path or EXPORTS_DIR / "Template_Impor_Buku.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Buku"
    ws.append(BUKU_HEADERS)
    ws.append(
        [
            "B0001",
            "Contoh Buku",
            "Penulis A",
            "Penerbit X",
            2024,
            "000",
            "Umum",
            "9786020000000",
            1,
            "BOS",
            50000,
            "id",
            "A1",
            "",
        ]
    )
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
def _read_rows(path: Path | str, headers: list[str], sheet: str | None = None) -> list[dict]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    file_headers = [str(c).strip() if c is not None else "" for c in rows[0]]

    out: list[dict] = []
    for raw in rows[1:]:
        rec: dict[str, object] = {}
        for idx, val in enumerate(raw):
            if idx >= len(file_headers):
                break
            key = file_headers[idx]
            if key in headers:
                rec[key] = val
        if any(v not in (None, "") for v in rec.values()):
            out.append(rec)
    return out


def import_anggota(
    path: Path | str, *, db: Database | None = None
) -> dict[str, int]:
    db = db or get_db()
    rows = _read_rows(path, ANGGOTA_HEADERS, sheet="Anggota")
    inserted = 0
    skipped = 0
    duplicates = 0
    for r in rows:
        try:
            if r.get("kode_anggota") and anggota_repo.get_by_kode(str(r["kode_anggota"]), db=db):
                duplicates += 1
                continue
            anggota_repo.create(
                {k: (str(v) if v is not None else None) for k, v in r.items()},
                db=db,
            )
            inserted += 1
        except Exception:
            skipped += 1
    return {"inserted": inserted, "skipped": skipped, "duplicates": duplicates, "total": len(rows)}


def import_buku(path: Path | str, *, db: Database | None = None) -> dict[str, int]:
    db = db or get_db()
    rows = _read_rows(path, BUKU_HEADERS, sheet="Buku")
    inserted = 0
    skipped = 0
    duplicates = 0
    for r in rows:
        try:
            if r.get("kode_buku") and buku_repo.get_by_kode(str(r["kode_buku"]), db=db):
                duplicates += 1
                continue
            buku_repo.create(
                {k: (v if v is not None else None) for k, v in r.items()}, db=db
            )
            inserted += 1
        except Exception:
            skipped += 1
    return {"inserted": inserted, "skipped": skipped, "duplicates": duplicates, "total": len(rows)}


# ---------------------------------------------------------------------------
# Export workbook (semua data)
# ---------------------------------------------------------------------------
def export_all_workbook(path: Path | None = None, db: Database | None = None) -> Path:
    """Ekspor semua data ke .xlsx dengan satu sheet per entitas."""
    db = db or get_db()
    out = path or EXPORTS_DIR / "Backup_Perpustakaan.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    sources: list[tuple[str, str]] = [
        ("Anggota", "SELECT * FROM anggota ORDER BY kode_anggota"),
        ("Buku", "SELECT * FROM buku ORDER BY kode_buku"),
        ("Eksemplar", "SELECT * FROM eksemplar ORDER BY kode_eksemplar"),
        ("Peminjaman", "SELECT * FROM peminjaman ORDER BY tanggal_pinjam DESC"),
        (
            "Peminjaman_Item",
            "SELECT * FROM peminjaman_item ORDER BY peminjaman_id DESC",
        ),
        ("Kunjungan", "SELECT * FROM kunjungan ORDER BY tanggal DESC"),
        ("Kas", "SELECT * FROM kas ORDER BY tanggal DESC"),
        ("Settings", "SELECT * FROM settings"),
    ]
    for sheet_name, sql in sources:
        ws = wb.create_sheet(sheet_name)
        rows = db.query_all(sql)
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h) for h in headers])
        else:
            ws.append(["(empty)"])
    wb.save(out)
    return out
